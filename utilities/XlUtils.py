import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column


def read_data(file, sheet_name, rowNum, colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=rowNum, column=colNum).value


def write_data(file, sheet_name, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(file)


